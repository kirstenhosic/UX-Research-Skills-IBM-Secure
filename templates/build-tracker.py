#!/usr/bin/env python3
"""
Build templates/06-participant-tracker.xlsx.

This script is the source of truth for the tracker; the .xlsx is generated.
After editing, re-run:  python3 templates/build-tracker.py
Requires openpyxl.

Every sheet gets the shared print furniture: a "Secure UX Design" header and
a right-aligned page number in the footer.

Part of the Dr. Morgan UX research suite.
Author: Kirsten Hosic, UX Research Strategy Lead, Security Product Design.
License: MIT.
"""
import os

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                   "06-participant-tracker.xlsx")

FONT = "IBM Plex Sans"
BLUE60 = "0F62FE"      # Carbon Blue 60
GRAY100 = "161616"     # Carbon Gray 100
GRAY70 = "525252"      # Carbon Gray 70
GRAY10 = "F4F4F4"      # Carbon Gray 10

header_font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
header_fill = PatternFill("solid", fgColor=BLUE60)
body_font = Font(name=FONT, size=10, color=GRAY100)
example_font = Font(name=FONT, size=10, color=GRAY70, italic=True)
wrap = Alignment(wrap_text=True, vertical="top")
thin = Side(style="thin", color="D0D0D0")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()

# ---------------------------------------------------------------- How to use
ws = wb.active
ws.title = "How to use"
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 4
ws.column_dimensions["B"].width = 100

rows = [
    ("title", "Participant tracker — [Study name]"),
    ("blank", ""),
    ("h", "What this is"),
    ("p", "One tracker per study, created during planning (research plan section 10), before the "
          "first recruit. It is the single consistent record of who is in the study, their consent "
          "terms, session logistics, and whether the loop was closed with them afterward."),
    ("blank", ""),
    ("h", "The privacy rule this file exists to enforce"),
    ("p", "This tracker is the ONLY place a participant's identity meets their P-ID. Everything "
          "released (findings records, report, deck, impact message) uses P-IDs only. Therefore: "
          "restrict access to the study team, store it in the location named in plan section 14, "
          "never attach or link it in a released artifact, and apply the study's retention period "
          "to it. If consent terms change, update the consent columns here first; downstream "
          "artifacts depend on them."),
    ("blank", ""),
    ("h", "How to fill it in"),
    ("p", "1. Assign P-IDs (P1, P2, ...) when a candidate enters from the screener. Sequential, "
          "never reused, never reassigned, even if the person withdraws."),
    ("p", "2. Internal participants and external participants live on their own tabs. External "
          "means customers or external SMEs; internal means IBM colleagues (direct users or "
          "proxies). The participant type column on each tab uses the suite's taxonomy: "
          "customer-direct / sme-external (external tab), internal-direct / internal-proxy "
          "(internal tab). This label travels onto every evidence entry in the findings records "
          "and sets the tier of the impact message."),
    ("p", "3. The consent columns (recording / quoting / recontact) come from the signed consent "
          "form. \"Quoting\" gates verbatim quotes in released artifacts; \"recontact\" gates the "
          "participant impact message. When in doubt, check the form itself; never guess."),
    ("p", "4. After the findings release, record the impact message date per participant. External "
          "and internal recipients get different drafts (see templates/10-participant-impact-message.md); "
          "send individually, never a visible CC list."),
    ("blank", ""),
    ("h", "Legend"),
    ("p", "Fill every white cell that applies. The gray italic row on each tab is an example of "
          "the expected format; delete it before real use. Dropdown columns (Status, Screener "
          "result, Participant type, consent columns) offer their allowed values when you select "
          "a cell."),
    ("blank", ""),
    ("p", "Part of the Dr. Morgan UX research suite (templates/). Author: Kirsten Hosic."),
]
r = 1
for kind, text in rows:
    c = ws.cell(row=r, column=2, value=text)
    if kind == "title":
        c.font = Font(name=FONT, bold=True, size=16, color=BLUE60)
    elif kind == "h":
        c.font = Font(name=FONT, bold=True, size=11, color=BLUE60)
    else:
        c.font = body_font
        c.alignment = wrap
        if text:
            # rough height for wrapped text at ~100 chars/line
            ws.row_dimensions[r].height = 14 * (len(text) // 100 + 1) + 4
    r += 1

# ------------------------------------------------------------- shared columns
def sheet(ws, cols, widths, example, dropdowns):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"
    for i, (name, w) in enumerate(zip(cols, widths), start=1):
        c = ws.cell(row=1, column=i, value=name)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(wrap_text=True, vertical="center")
        c.border = border
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 30
    for i, v in enumerate(example, start=1):
        c = ws.cell(row=2, column=i, value=v)
        c.font = example_font
        c.border = border
        c.alignment = wrap
    # style blank data rows
    for row in range(3, 33):
        for col in range(1, len(cols) + 1):
            c = ws.cell(row=row, column=col)
            c.font = body_font
            c.border = border
            if row % 2 == 1:
                c.fill = PatternFill("solid", fgColor=GRAY10)
    for col_name, values in dropdowns.items():
        idx = cols.index(col_name) + 1
        letter = get_column_letter(idx)
        dv = DataValidation(type="list", formula1='"' + ",".join(values) + '"',
                            allow_blank=True, showErrorMessage=True)
        dv.error = "Pick a value from the list."
        ws.add_data_validation(dv)
        dv.add(f"{letter}2:{letter}200")

YN = ["Yes", "No", "Unknown"]
STATUS = ["Candidate", "Screened", "Scheduled", "Completed", "No-show", "Withdrawn", "Excluded"]
SCREEN = ["Pending", "Qualified", "Disqualified"]

# ------------------------------------------------------------------- Internal
wsi = wb.create_sheet("Internal participants")
cols_i = ["P-ID", "Status", "Name", "Email", "Business unit / team", "Role / job title",
          "Participant type", "Persona", "Product", "Recruitment source", "Screener result",
          "Consent: recording", "Consent: quoting", "Consent: recontact", "Consent form location",
          "Session date", "Session time (TZ)", "Moderator", "Notetaker",
          "Transcript / notes location", "Impact message sent (date)", "Notes"]
widths_i = [7, 11, 18, 24, 20, 20, 15, 18, 14, 18, 12,
            10, 10, 10, 24, 12, 13, 14, 14, 26, 14, 30]
example_i = ["P1", "Completed", "Ada Example", "ada.example@ibm.com", "Secure Platform / SRE",
             "Site reliability engineer", "internal-direct", "Platform operator", "Vault",
             "Slack #vault-users call for participants", "Qualified",
             "Yes", "Yes", "Yes", "Box > Study folder > consent/P1.pdf",
             "2026-09-15", "10:00 ET", "K. Hosic", "J. Doe",
             "Box > Study folder > transcripts/transcript-p1.txt", "2026-10-02",
             "Example row — delete before use"]
sheet(wsi, cols_i, widths_i, example_i, {
    "Status": STATUS, "Screener result": SCREEN,
    "Participant type": ["internal-direct", "internal-proxy"],
    "Consent: recording": YN, "Consent: quoting": YN, "Consent: recontact": YN,
})

# ------------------------------------------------------------------- External
wse = wb.create_sheet("External participants")
cols_e = ["P-ID", "Status", "Name", "Email", "Company / organization", "Role / job title",
          "Participant type", "Persona", "Product", "Routed via (PM / account contact)",
          "Recruitment source", "Screener result", "NDA / consent form location",
          "Consent: recording", "Consent: quoting", "Consent: recontact",
          "Session date", "Session time (TZ)", "Moderator", "Notetaker",
          "Incentive (type / amount)", "Incentive sent (date)",
          "Transcript / notes location", "Impact message sent (date)", "Notes"]
widths_e = [7, 11, 18, 24, 22, 20, 15, 18, 14, 24, 18, 12, 26,
            10, 10, 10, 12, 13, 14, 14, 18, 13, 26, 14, 30]
example_e = ["P2", "Scheduled", "Sam Example", "sam@example.com", "Example Corp",
             "Platform engineering lead", "customer-direct", "Platform operator", "Vault",
             "A. PM (Vault account team)", "Account team outreach email", "Qualified",
             "Box > Study folder > consent/P2.pdf", "Yes", "No", "Yes",
             "2026-09-18", "13:00 ET", "K. Hosic", "J. Doe",
             "Gift card / $100", "", "Box > Study folder > transcripts/transcript-p2.txt", "",
             "Example row — delete before use. Quoting consent is No: paraphrase only."]
sheet(wse, cols_e, widths_e, example_e, {
    "Status": STATUS, "Screener result": SCREEN,
    "Participant type": ["customer-direct", "sme-external"],
    "Consent: recording": YN, "Consent: quoting": YN, "Consent: recontact": YN,
})

# Print furniture on every sheet: header left, page number right in footer.
for sheet_ws in wb:
    sheet_ws.oddHeader.left.text = "Secure UX Design"
    sheet_ws.oddHeader.left.font = f"{FONT},Bold"
    sheet_ws.oddHeader.left.size = 9
    sheet_ws.oddFooter.right.text = "Page &P of &N"
    sheet_ws.oddFooter.right.font = FONT
    sheet_ws.oddFooter.right.size = 9

# Deterministic output: pinned document properties and zip timestamps, so a
# no-change rebuild leaves git status clean (same bar as build-skill.sh).
# openpyxl refreshes dcterms:modified at save time, so it is pinned in the
# rewrite below rather than on wb.properties.
from datetime import datetime
import re
import zipfile
wb.properties.created = datetime(2026, 9, 2)
wb.save(OUT)
tmp = OUT + ".tmp"
with zipfile.ZipFile(OUT) as zin, \
        zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
    for name in sorted(zin.namelist()):
        data = zin.read(name)
        if name == "docProps/core.xml":
            data = re.sub(
                rb"(<dcterms:modified[^>]*>)[^<]*(</dcterms:modified>)",
                rb"\g<1>2026-09-02T00:00:00Z\g<2>", data)
        info = zipfile.ZipInfo(name, date_time=(1980, 1, 1, 0, 0, 0))
        info.compress_type = zipfile.ZIP_DEFLATED
        zout.writestr(info, data)
os.replace(tmp, OUT)
print("saved", OUT)
